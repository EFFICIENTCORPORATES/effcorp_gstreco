from tkinter import *
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import os
import glob
from openpyxl import load_workbook
from shutil import copyfile
import datetime
import warnings
import numpy as np



warnings.filterwarnings('ignore')


#enter the path of the GSTR2A file
gstr2apath=input("Enter the complete path of the GSTR2A Excel File: ")

gstr2a=pd.read_excel(str(gstr2apath), sheet_name="Main_2A_Format",dtype={"Inv_CN_DN_Number_Final":str, "Inv_CN_DN_Date_Text":str, "Total_Tax":int})


#enter the path of the Purchase Register File

itrpath=input("Enter the complete path of the Purchase Register: ")

itr=pd.read_excel(str(itrpath), sheet_name="Main_ITR_Format",dtype={"Invoice_Number":str, "Invoice_Date_Text":str,"Total_Tax":int})



pth = os.path.dirname(gstr2apath)

fullpath1 = pth + "/" + "Workings.xlsx"

writer = pd.ExcelWriter(fullpath1, engine='xlsxwriter', options={'strings_to_formulas': True})


fullpath1a = pth + "/" + "Summary.xlsx"
writer1 = pd.ExcelWriter(fullpath1a, engine='xlsxwriter', options={'strings_to_formulas': True})

df1 = pd.DataFrame()
df1.to_excel(writer1, sheet_name="Summary", index=False)

writer1.save()

fullpath2 = fullpath1a.replace("/", "\\")  # this is a  useful command for defining the correct filepath

wb = load_workbook(fullpath2)
ws = wb["Summary"]

ws["B2"].value = "SUMMARY OF THE RECONCILIATION OF GSTR2A Vs ITR"
ws.merge_cells("B2:F2")
ws["C4"].value = "GSTR2A"
ws.merge_cells("C4:D4")
ws["E4"].value = "Purchase Register"
ws.merge_cells("E4:F4")

ws["B4"].value = "Particulars"
ws.merge_cells("B4:B5")
ws["C5"].value = "Count"
ws["D5"].value = "Tax Amount"
ws["E5"].value = "Count"
ws["F5"].value = "Tax Amount"
ws["B7"].value = "Total To be Matched"
ws["B9"].value = "Matched with GST_INVNO_INVDATE_3_WAY"
ws["B10"].value = "Matched with GST_INVNO_2_WAY"
ws["B11"].value = "Matched with GST_INVDATE_2_WAY"

ws["B13"].value = "Identified Possible Matches"

ws["B15"].value = "Matched with PAN_INVNO_INVDATE_3_WAY"
ws["B16"].value = "Matched with PAN_INVNO_2_WAY"
ws["B17"].value = "Matched with PAN_INVDATE_2_WAY"

ws["B19"].value = "Unmatched Cases"

ws["B21"].value = "Check"

# setting the tolerance limit for matching in Rupees

tol_limit = int(150)

ws["F1"].value = f"Tolerance Limit was {tol_limit}"

try:
    gstr2a['Inv_CN_DN_Number_Finall'] = gstr2a["Inv_CN_DN_Number_Final"].apply(lambda x: x.lower(str()))
except:
    gstr2a['Inv_CN_DN_Number_Finall'] = gstr2a["Inv_CN_DN_Number_Final"]

gstr2a['GST_INVNO_INVDATE_3_WAY'] = gstr2a['GSTIN_of_Supplier'] + "/" + gstr2a['Inv_CN_DN_Number_Finall'] + "/" + gstr2a['Inv_CN_DN_Date_Text']

gstr2a['GST_INVNO_2_WAY'] = gstr2a['GSTIN_of_Supplier'] + "/" + gstr2a['Inv_CN_DN_Number_Finall']

gstr2a['GST_INVDATE_2_WAY'] = gstr2a['GSTIN_of_Supplier'] + "/" + gstr2a['Inv_CN_DN_Date_Text']

try:
    gstr2a['PAN_Number'] = gstr2a["GSTIN_of_Supplier"].apply(lambda x: x[2:12:1])
except:
    gstr2a['PAN_Number'] = gstr2a["GSTIN_of_Supplier"]

# the PAN number matches will be used as possible matches

gstr2a['PAN_INVNO_INVDATE_3_WAY'] = gstr2a['PAN_Number'] + "/" + gstr2a['Inv_CN_DN_Number_Finall'] + "/" + \
                                    gstr2a['Inv_CN_DN_Date_Text']

gstr2a['PAN_INVNO_2_WAY'] = gstr2a['PAN_Number'] + "/" + gstr2a['Inv_CN_DN_Number_Finall']

gstr2a['PAN_INVDATE_2_WAY'] = gstr2a['PAN_Number'] + "/" + gstr2a['Inv_CN_DN_Date_Text']


try:
    itr["Invoice_Numberl"] = itr["Invoice_Number"].apply(lambda x: x.lower(str()))
except:
    itr["Invoice_Numberl"] = itr["Invoice_Number"]

itr["GST_INVNO_INVDATE_3_WAY"] = itr["Vendor_GST_REG"] + "/" + itr["Invoice_Numberl"] + "/" + itr[
    "Invoice_Date_Text"]

itr["GST_INVNO_2_WAY"] = itr["Vendor_GST_REG"] + "/" + itr["Invoice_Numberl"]

itr["GST_INVDATE_2_WAY"] = itr["Vendor_GST_REG"] + "/" + itr["Invoice_Date_Text"]

try:
    itr["PAN_Number"] = itr["Vendor_GST_REG"].apply(lambda x: x[2:12:1])
except:
    itr["PAN_Number"] = itr["Vendor_GST_REG"]

# the PAN number matches will be used as possible matches

itr["PAN_INVNO_INVDATE_3_WAY"] = itr["PAN_Number"] + "/" + itr["Invoice_Numberl"] + "/" + itr["Invoice_Date_Text"]

itr["PAN_INVNO_2_WAY"] = itr["PAN_Number"] + "/" + itr["Invoice_Numberl"]

itr["PAN_INVDATE_2_WAY"] = itr["PAN_Number"] + "/" + itr["Invoice_Date_Text"]

ws["C7"].value = list(gstr2a.shape)[0]
ws["D7"].value = sum(gstr2a["Total_Tax"])
ws["E7"].value = list(itr.shape)[0]
ws["F7"].value = sum(itr["Total_Tax"])

# First Cut Matching : Here we will try to do that Matching based on 3 way i.e GST No, Inv No & Inv Date being same in ITR & GSTR2A

gstr2a_pivot = pd.pivot_table(gstr2a, values="Total_Tax", index=["GST_INVNO_INVDATE_3_WAY"], aggfunc=np.sum)

itr_pivot = pd.pivot_table(itr, values="Total_Tax", index=["GST_INVNO_INVDATE_3_WAY"], aggfunc=np.sum)

gstr2a_pivot.rename(columns={'Total_Tax': 'Tax_as_per_GSTR2A'}, inplace=True)

itr_pivot.rename(columns={'Total_Tax': 'Tax_as_per_ITR'}, inplace=True)

gstr2a_pivot.reset_index(inplace=True)

itr_pivot.reset_index(inplace=True)

compared = gstr2a_pivot.merge(itr_pivot, left_on="GST_INVNO_INVDATE_3_WAY", right_on="GST_INVNO_INVDATE_3_WAY",
                              how="left")

compared = compared.replace(np.nan, 0, regex=True)

compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

conditions = [compared["Difference_in_Tax"] > (tol_limit),

              compared["Difference_in_Tax"] < (tol_limit * -1),

              ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

              ]

results = ["Excess in GSTR 2A, Less in ITR",

           "Excess in ITR, Less in GSTR2A",

           "Exact Match within Tolerance"]

compared["Remarks"] = np.select(conditions, results)

# The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

# now we will select the exact match within the Tolerance level

mask_1 = compared["Remarks"].values == "Exact Match within Tolerance"

match_3_way_list = compared[mask_1]["GST_INVNO_INVDATE_3_WAY"].values

mask_1a = gstr2a["GST_INVNO_INVDATE_3_WAY"].isin(match_3_way_list)  # returns a Boolean Array

mask_1b = itr["GST_INVNO_INVDATE_3_WAY"].isin(match_3_way_list)  # returns a boolean array

matched_gstr2a_3way = gstr2a[mask_1a]
matched_gstr2a_3way["Matching Category"] = "3 Way matching GST + Inv No+ Inv Date"
matched_itr_3way = itr[mask_1b]
matched_itr_3way["Matching Category"] = "3 Way matching GST + Inv No+ Inv Date"

ws["C9"].value = len(matched_gstr2a_3way["GST_INVNO_INVDATE_3_WAY"])
ws["D9"].value = sum(matched_gstr2a_3way["Total_Tax"])
ws["E9"].value = len(matched_itr_3way["GST_INVNO_INVDATE_3_WAY"])
ws["F9"].value = sum(matched_itr_3way["Total_Tax"])

bal_gstr2a_1cut = gstr2a[~mask_1a]
bal_itr_1cut = itr[~mask_1b]

# Second Cut Matching : Here we will try to do that Matching based on 2 way i.e GST No & Inv No

gstr2a_pivot = pd.pivot_table(bal_gstr2a_1cut, values="Total_Tax", index=["GST_INVNO_2_WAY"], aggfunc=np.sum)

itr_pivot = pd.pivot_table(bal_itr_1cut, values="Total_Tax", index=["GST_INVNO_2_WAY"], aggfunc=np.sum)

gstr2a_pivot.rename(columns={'Total_Tax': 'Tax_as_per_GSTR2A'}, inplace=True)

itr_pivot.rename(columns={'Total_Tax': 'Tax_as_per_ITR'}, inplace=True)

gstr2a_pivot.reset_index(inplace=True)

itr_pivot.reset_index(inplace=True)

compared = gstr2a_pivot.merge(itr_pivot, left_on="GST_INVNO_2_WAY", right_on="GST_INVNO_2_WAY", how="left")

compared = compared.replace(np.nan, 0, regex=True)

compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

conditions = [compared["Difference_in_Tax"] > (tol_limit),

              compared["Difference_in_Tax"] < (tol_limit * -1),

              ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

              ]

results = ["Excess in GSTR 2A, Less in ITR",

           "Excess in ITR, Less in GSTR2A",

           "Exact Match within Tolerance"]

compared["Remarks"] = np.select(conditions, results)

# The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

# now we will select the exact match within the Tolerance level

mask_1 = compared["Remarks"].values == "Exact Match within Tolerance"

match_2_way_list1 = compared[mask_1]["GST_INVNO_2_WAY"].values

mask_1a = bal_gstr2a_1cut["GST_INVNO_2_WAY"].isin(match_2_way_list1)  # returns a Boolean Array

mask_1b = bal_itr_1cut["GST_INVNO_2_WAY"].isin(match_2_way_list1)  # returns a boolean array

matched_gstr2a_2way1 = bal_gstr2a_1cut[mask_1a]
matched_itr_2way1 = bal_itr_1cut[mask_1b]

matched_gstr2a_2way1["Matching Category"] = "2 Way matching GST + Inv No"
matched_itr_2way1["Matching Category"] = "2 Way matching GST + Inv No"

ws["C10"].value = len(matched_gstr2a_2way1["GST_INVNO_2_WAY"])
ws["D10"].value = sum(matched_gstr2a_2way1["Total_Tax"])
ws["E10"].value = len(matched_itr_2way1["GST_INVNO_2_WAY"])
ws["F10"].value = sum(matched_itr_2way1["Total_Tax"])

bal_gstr2a_2cut = bal_gstr2a_1cut[~mask_1a]
bal_itr_2cut = bal_itr_1cut[~mask_1b]

# Third Cut Matching : Here we will try to do that Matching based on 2 way i.e GST No & Inv Date

gstr2a_pivot = pd.pivot_table(bal_gstr2a_2cut, values="Total_Tax", index=["GST_INVDATE_2_WAY"], aggfunc=np.sum)

itr_pivot = pd.pivot_table(bal_itr_2cut, values="Total_Tax", index=["GST_INVDATE_2_WAY"], aggfunc=np.sum)

gstr2a_pivot.rename(columns={'Total_Tax': 'Tax_as_per_GSTR2A'}, inplace=True)

itr_pivot.rename(columns={'Total_Tax': 'Tax_as_per_ITR'}, inplace=True)

gstr2a_pivot.reset_index(inplace=True)

itr_pivot.reset_index(inplace=True)

compared = gstr2a_pivot.merge(itr_pivot, left_on="GST_INVDATE_2_WAY", right_on="GST_INVDATE_2_WAY", how="left")

compared = compared.replace(np.nan, 0, regex=True)

compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

conditions = [compared["Difference_in_Tax"] > (tol_limit),

              compared["Difference_in_Tax"] < (tol_limit * -1),

              ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

              ]

results = ["Excess in GSTR 2A, Less in ITR",

           "Excess in ITR, Less in GSTR2A",

           "Exact Match within Tolerance"]

compared["Remarks"] = np.select(conditions, results)

# The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

# now we will select the exact match within the Tolerance level

mask_1 = compared["Remarks"].values == "Exact Match within Tolerance"

match_2_way_list2 = compared[mask_1]["GST_INVDATE_2_WAY"].values

mask_1a = bal_gstr2a_2cut["GST_INVDATE_2_WAY"].isin(match_2_way_list2)  # returns a Boolean Array

mask_1b = bal_itr_2cut["GST_INVDATE_2_WAY"].isin(match_2_way_list2)  # returns a boolean array

matched_gstr2a_2way2 = bal_gstr2a_2cut[mask_1a]
matched_itr_2way2 = bal_itr_2cut[mask_1b]

matched_gstr2a_2way2["Matching Category"] = "2 Way matching GST + Inv Date"
matched_itr_2way2["Matching Category"] = "2 Way matching GST + Inv Date"

ws["C11"].value = len(matched_gstr2a_2way2["GST_INVDATE_2_WAY"])
ws["D11"].value = sum(matched_gstr2a_2way2["Total_Tax"])
ws["E11"].value = len(matched_itr_2way2["GST_INVDATE_2_WAY"])
ws["F11"].value = sum(matched_itr_2way2["Total_Tax"])

bal_gstr2a_3cut = bal_gstr2a_2cut[~mask_1a]
bal_itr_3cut = bal_itr_2cut[~mask_1b]

#after the 3 cut matching, now we try to find out the Possible matches in form of PAN matching and upper /lower case matching
# Fourth Cut Matching : Here we will try to do that Matching based on 3 way With PAN No Inv NO and Inv Date

gstr2a_pivot = pd.pivot_table(bal_gstr2a_3cut, values="Total_Tax", index=["PAN_INVNO_INVDATE_3_WAY"], aggfunc=np.sum)

itr_pivot = pd.pivot_table(bal_itr_3cut, values="Total_Tax", index=["PAN_INVNO_INVDATE_3_WAY"], aggfunc=np.sum)

gstr2a_pivot.rename(columns={'Total_Tax': 'Tax_as_per_GSTR2A'}, inplace=True)

itr_pivot.rename(columns={'Total_Tax': 'Tax_as_per_ITR'}, inplace=True)

gstr2a_pivot.reset_index(inplace=True)

itr_pivot.reset_index(inplace=True)

compared = gstr2a_pivot.merge(itr_pivot, left_on="PAN_INVNO_INVDATE_3_WAY", right_on="PAN_INVNO_INVDATE_3_WAY", how="left")

compared = compared.replace(np.nan, 0, regex=True)

compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

conditions = [compared["Difference_in_Tax"] > (tol_limit),

              compared["Difference_in_Tax"] < (tol_limit * -1),

              ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

              ]

results = ["Excess in GSTR 2A, Less in ITR",

           "Excess in ITR, Less in GSTR2A",

           "Exact Match within Tolerance"]

compared["Remarks"] = np.select(conditions, results)

# The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

# now we will select the exact match within the Tolerance level

mask_1 = compared["Remarks"].values == "Exact Match within Tolerance"

match_3_way_list2 = compared[mask_1]["PAN_INVNO_INVDATE_3_WAY"].values

mask_1a = bal_gstr2a_3cut["PAN_INVNO_INVDATE_3_WAY"].isin(match_3_way_list2)  # returns a Boolean Array

mask_1b = bal_itr_3cut["PAN_INVNO_INVDATE_3_WAY"].isin(match_3_way_list2)  # returns a boolean array

matched_gstr2a_3way2 = bal_gstr2a_3cut[mask_1a]
matched_itr_3way2 = bal_itr_3cut[mask_1b]

matched_gstr2a_3way2["Matching Category"] = "3 Way matching PAN + Inv No+ Inv Date"
matched_itr_3way2["Matching Category"] = "3 Way matching PAN + Inv No + Inv Date"

ws["C15"].value = len(matched_gstr2a_3way2["PAN_INVNO_INVDATE_3_WAY"])
ws["D15"].value = sum(matched_gstr2a_3way2["Total_Tax"])
ws["E15"].value = len(matched_itr_3way2["PAN_INVNO_INVDATE_3_WAY"])
ws["F15"].value = sum(matched_itr_3way2["Total_Tax"])

bal_gstr2a_4cut = bal_gstr2a_3cut[~mask_1a]
bal_itr_4cut = bal_itr_3cut[~mask_1b]

# Fifth Cut Matching : Here we will try to do that Matching based on 3 way With PAN No Inv NO and Inv Date

gstr2a_pivot = pd.pivot_table(bal_gstr2a_4cut, values="Total_Tax", index=["PAN_INVNO_2_WAY"],
                              aggfunc=np.sum)

itr_pivot = pd.pivot_table(bal_itr_4cut, values="Total_Tax", index=["PAN_INVNO_2_WAY"], aggfunc=np.sum)

gstr2a_pivot.rename(columns={'Total_Tax': 'Tax_as_per_GSTR2A'}, inplace=True)

itr_pivot.rename(columns={'Total_Tax': 'Tax_as_per_ITR'}, inplace=True)

gstr2a_pivot.reset_index(inplace=True)

itr_pivot.reset_index(inplace=True)

compared = gstr2a_pivot.merge(itr_pivot, left_on="PAN_INVNO_2_WAY", right_on="PAN_INVNO_2_WAY",
                              how="left")

compared = compared.replace(np.nan, 0, regex=True)

compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

conditions = [compared["Difference_in_Tax"] > (tol_limit),

              compared["Difference_in_Tax"] < (tol_limit * -1),

              ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

              ]

results = ["Excess in GSTR 2A, Less in ITR",

           "Excess in ITR, Less in GSTR2A",

           "Exact Match within Tolerance"]

compared["Remarks"] = np.select(conditions, results)

# The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

# now we will select the exact match within the Tolerance level

mask_1 = compared["Remarks"].values == "Exact Match within Tolerance"

match_2_way_list3 = compared[mask_1]["PAN_INVNO_2_WAY"].values

mask_1a = bal_gstr2a_4cut["PAN_INVNO_2_WAY"].isin(match_2_way_list3)  # returns a Boolean Array

mask_1b = bal_itr_4cut["PAN_INVNO_2_WAY"].isin(match_2_way_list3)  # returns a boolean array

matched_gstr2a_2way3 = bal_gstr2a_4cut[mask_1a]
matched_itr_2way3 = bal_itr_4cut[mask_1b]

matched_gstr2a_2way3["Matching Category"] = "2 Way matching PAN + Inv No"
matched_itr_2way3["Matching Category"] = "2 Way matching PAN + Inv No "

ws["C16"].value = len(matched_gstr2a_2way3["PAN_INVNO_2_WAY"])
ws["D16"].value = sum(matched_gstr2a_2way3["Total_Tax"])
ws["E16"].value = len(matched_itr_2way3["PAN_INVNO_2_WAY"])
ws["F16"].value = sum(matched_itr_2way3["Total_Tax"])

bal_gstr2a_5cut = bal_gstr2a_4cut[~mask_1a]
bal_itr_5cut = bal_itr_4cut[~mask_1b]



# Sixth Cut Matching : Here we will try to do that Matching based on 3 way With PAN No Inv NO and Inv Date

gstr2a_pivot = pd.pivot_table(bal_gstr2a_5cut, values="Total_Tax", index=["PAN_INVDATE_2_WAY"],
                              aggfunc=np.sum)

itr_pivot = pd.pivot_table(bal_itr_4cut, values="Total_Tax", index=["PAN_INVDATE_2_WAY"], aggfunc=np.sum)

gstr2a_pivot.rename(columns={'Total_Tax': 'Tax_as_per_GSTR2A'}, inplace=True)

itr_pivot.rename(columns={'Total_Tax': 'Tax_as_per_ITR'}, inplace=True)

gstr2a_pivot.reset_index(inplace=True)

itr_pivot.reset_index(inplace=True)

compared = gstr2a_pivot.merge(itr_pivot, left_on="PAN_INVDATE_2_WAY", right_on="PAN_INVDATE_2_WAY",
                              how="left")

compared = compared.replace(np.nan, 0, regex=True)

compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

conditions = [compared["Difference_in_Tax"] > (tol_limit),

              compared["Difference_in_Tax"] < (tol_limit * -1),

              ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

              ]

results = ["Excess in GSTR 2A, Less in ITR",

           "Excess in ITR, Less in GSTR2A",

           "Exact Match within Tolerance"]

compared["Remarks"] = np.select(conditions, results)

# The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

# now we will select the exact match within the Tolerance level

mask_1 = compared["Remarks"].values == "Exact Match within Tolerance"

match_2_way_list4 = compared[mask_1]["PAN_INVDATE_2_WAY"].values

mask_1a = bal_gstr2a_5cut["PAN_INVDATE_2_WAY"].isin(match_2_way_list4)  # returns a Boolean Array

mask_1b = bal_itr_5cut["PAN_INVDATE_2_WAY"].isin(match_2_way_list4)  # returns a boolean array

matched_gstr2a_2way4 = bal_gstr2a_5cut[mask_1a]
matched_itr_2way4 = bal_itr_5cut[mask_1b]

matched_gstr2a_2way4["Matching Category"] = "2 Way matching PAN + Inv Date"
matched_itr_2way4["Matching Category"] = "2 Way matching PAN + Inv Date "

ws["C17"].value = len(matched_gstr2a_2way4["PAN_INVDATE_2_WAY"])
ws["D17"].value = sum(matched_gstr2a_2way4["Total_Tax"])
ws["E17"].value = len(matched_itr_2way4["PAN_INVDATE_2_WAY"])
ws["F17"].value = sum(matched_itr_2way4["Total_Tax"])

bal_gstr2a_6cut = bal_gstr2a_5cut[~mask_1a]
bal_itr_6cut = bal_itr_5cut[~mask_1b]

gstr2a.to_excel(writer, sheet_name='Orignal GSTR2A', index=False)

itr.to_excel(writer, sheet_name='Original ITR', index=False)

all_matched_2a = pd.concat([matched_gstr2a_3way, matched_gstr2a_2way1, matched_gstr2a_2way2,matched_gstr2a_3way2, matched_gstr2a_2way3,matched_gstr2a_2way4], ignore_index=True)

all_matched_itr = pd.concat([matched_itr_3way, matched_itr_2way1, matched_itr_2way2,matched_itr_3way2,matched_itr_2way3, matched_itr_2way4], ignore_index=True)

all_matched_2a.to_excel(writer, sheet_name='Matched_GSTR2A', index=False)

all_matched_itr.to_excel(writer, sheet_name='Matched_ITR', index=False)

bal_gstr2a_6cut.to_excel(writer, sheet_name='Unmatched_GSTR2A', index=False)

bal_itr_6cut.to_excel(writer, sheet_name='Unmatched_ITR', index=False)

ws["C19"].value = len(bal_gstr2a_6cut["GST_INVDATE_2_WAY"])
ws["D19"].value = sum(bal_gstr2a_6cut["Total_Tax"])
ws["E19"].value = len(bal_itr_6cut["GST_INVDATE_2_WAY"])
ws["F19"].value = sum(bal_itr_6cut["Total_Tax"])

writer.save()


wb.save(fullpath2)
writer.save()

print(f"The Reconciliation has been Done and files are stored in {fullpath2}")