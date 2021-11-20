from tkinter import *
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import os
import glob
from openpyxl import load_workbook
from shutil import copyfile
import datetime
import warnings


FotaGui = Tk()
LogGui = Tk()

FotaGui.geometry('600x700')
FotaGui.title('GSTR2A Reconciliation for Corporates  ')

LogGui.geometry('800x400')
LogGui.title('Log of all activities:')

filepath = ''

warnings.filterwarnings('ignore')


def HELP_INFO():
    root = Tk()

    label_head0 = Label(root, text="   Program for GSTR2A Reconciliation", font="Times 10 ")

    label_head0.pack()

    label_head1 = Label(root, text='For a video guide on usage of this Program , Please visit YT Channel Efficient Corporates .'
                                   ' \n   .', bd=1, relief='solid', font='Times 13', anchor=N)
    label_head1.pack()

    label_head6 = Label(root, text='For Any Other Queries/ Issues/ Feedback: Please Contact'
                                   '\n Team Efficient Corporates'
                                   '\n Email= efficientcorporates.info@gmail.com'
                                   '\n Mob: 7464098001'
                                   '\n'
                                   '\n We want feedbak on this program so that it can be further improved and enhabced to meet needs of more users',
                        bd=1, relief='solid',
                        font='Times 11', anchor=N)
    label_head6.pack()

    root.title('Help/ Info About the Program')

    root.mainloop()


def file_path_mon2a():
    global filepath
    global files_itr
    global files_con2a
    global files_mon2a

    global filepath
    global label_head7
    filepath = StringVar()
    global now

    now = datetime.datetime.now()

    # Fetch the file path of the hex file browsed.
    if (filepath == ""):
        filepath = filedialog.askopenfilename(initialdir=os.getcwd(),
                                              title="Select a file",
                                              filetypes=[("All Files", "*.*"), ("Pdf Files", "*.pdf"),
                                                         ("Text Files", "*.txt"), ("Excel FIles", "*.xlsx")])
    else:
        filepath = filedialog.askopenfilename(initialdir=filepath,
                                              title="Select a file",
                                              filetypes=[("All Files", "*.*"), ("Pdf Files", "*.pdf"),
                                                         ("Text Files", "*.txt"), ("Excel FIles", "*.xlsx")])

    extension = os.path.splitext(filepath)[1]
    filename = os.path.splitext(filepath)[0]
    pth = os.path.dirname(filepath)
    files_mon2a = glob.glob(os.path.join(pth, '*{ext}'.format(ext=extension)))

    for f in files_mon2a:
        label_head7 = Label(LogGui, text='{n}The File {fil} have been selected.'.format(fil=f, n=now.strftime(
            '%y-%m-%d %H:%M:%S')), bd=1, relief='solid',
                            font='Times 10', anchor=N)
        label_head7.pack()

    return files_mon2a


def file_path_con2a():
    global filepath
    global files_itr
    global files_con2a
    global files_mon2a

    global label_head7
    filepath = StringVar()
    global now

    now = datetime.datetime.now()

    # Fetch the file path of the hex file browsed.
    if (filepath == ""):
        filepath = filedialog.askopenfilename(initialdir=os.getcwd(),
                                              title="Select a file",
                                              filetypes=[("All Files", "*.*"), ("Pdf Files", "*.pdf"),
                                                         ("Text Files", "*.txt"), ("Excel FIles", "*.xlsx")])
    else:
        filepath = filedialog.askopenfilename(initialdir=filepath,
                                              title="Select a file",
                                              filetypes=[("All Files", "*.*"), ("Pdf Files", "*.pdf"),
                                                         ("Text Files", "*.txt"), ("Excel FIles", "*.xlsx")])

    extension = os.path.splitext(filepath)[1]
    filename = os.path.splitext(filepath)[0]
    pth = os.path.dirname(filepath)
    files_con2a = glob.glob(os.path.join(pth, '*{ext}'.format(ext=extension)))
    files_con2a = list(map(lambda st: str.replace(st, "\\", "/"), files_con2a))

    for f in files_con2a:
        label_head7 = Label(LogGui, text='{n}The File {fil} have been selected.'.format(fil=f, n=now.strftime(
            '%y-%m-%d %H:%M:%S')), bd=1, relief='solid',
                            font='Times 10', anchor=N)
        label_head7.pack()

    return files_con2a


def file_path_itr():
    global filepath
    global files_itr
    global files_itr
    global files_con2a
    global files_mon2a

    global label_head7
    filepath = StringVar()
    global now

    now = datetime.datetime.now()

    # Fetch the file path of the hex file browsed.
    if (filepath == ""):
        filepath = filedialog.askopenfilename(initialdir=os.getcwd(),
                                              title="Select a file",
                                              filetypes=[("All Files", "*.*"), ("Pdf Files", "*.pdf"),
                                                         ("Text Files", "*.txt"), ("Excel FIles", "*.xlsx")])
    else:
        filepath = filedialog.askopenfilename(initialdir=filepath,
                                              title="Select a file",
                                              filetypes=[("All Files", "*.*"), ("Pdf Files", "*.pdf"),
                                                         ("Text Files", "*.txt"), ("Excel FIles", "*.xlsx")])

    extension = os.path.splitext(filepath)[1]
    filename = os.path.splitext(filepath)[0]
    pth = os.path.dirname(filepath)
    files_itr = glob.glob(os.path.join(pth, '*{ext}'.format(ext=extension)))
    files_itr = list(map(lambda st: str.replace(st, "\\", "/"), files_itr))

    for f in files_itr:
        label_head7 = Label(LogGui, text='{n}The File {fil} have been selected.'.format(fil=f, n=now.strftime(
            '%y-%m-%d %H:%M:%S')), bd=1, relief='solid',
                            font='Times 10', anchor=N)
        label_head7.pack()

    return files_itr


def sendtofile(colslist, filepath):
    df = pd.read_excel(filepath)
    cols = e_1.get()
    pth = os.path.dirname(filepath)
    colslist = list(set(df[cols].values))
    global now

    for i in colslist:
        df[df[cols] == i].to_excel("{}/{}.xlsx".format(pth, i), sheet_name=i, index=False)

    messagebox.showinfo('Output',
                        'You data has been split into {} and {} files has been created.Click OK. \n All Files stored in same folder'.format(
                            ', '.join(colslist), len(colslist)))

    label_head7 = Label(LogGui,
                        text='{n}The Files have been Splitted to different Files.'.format(
                            n=now.strftime('%y-%m-%d %H:%M:%S')),
                        bd=1, relief='solid',
                        font='Times 10', anchor=N)
    label_head7.pack()

    print('\nCompleted')
    print('Thanks for using this program.')
    return


def SPLIT_FILE():
    global filepath
    global e_1
    global e_2

    splitwin = Tk()

    label_1 = Label(splitwin, text='Enter the Exact Column name whose value u want to Split')
    label_1.pack()
    e_1 = Entry(splitwin, width=50, bg='blue', fg='white', borderwidth=4)
    e_1.pack()

    Browsebutton = Button(splitwin, width=20, text="Split Files", command=SPLIT_FILE2)
    Browsebutton.pack()

    splitwin.mainloop()


def SPLIT_FILE2():
    df = pd.read_excel(filepath)
    cols = e_1.get()
    colslist = list(set(df[cols].values))

    messagebox.showinfo('Check the output',
                        'You data will split based on these values {} and create {} files or sheets based on next selection. If you are ready to proceed Click OK or close the dialog box to re-start.'.format(
                            ', '.join(colslist), len(colslist)))

    response = messagebox.askyesno('Split Files',
                                   'Do you want to split in Various Sheets in Same file  OR Different Files? '
                                   '\nClick Yes for Various Sheets in Same File.!'
                                   '\n CLick No For Different Files')
    df = pd.read_excel(filepath)
    cols = e_1.get()
    colslist = list(set(df[cols].values))

    if response == 0:
        sendtofile(colslist, filepath)
    elif response == 1:
        sendtosheet(colslist)
    else:
        messagebox.showerror('Output', "Something went wrong")


def sendtosheet(colslist):
    cols = e_1.get()
    extension = os.path.splitext(filepath)[1]
    filename = os.path.splitext(filepath)[0]
    pth = os.path.dirname(filepath)
    newfile = os.path.join(pth, filename + '_Sheet_Split_Auto' + extension)
    df = pd.read_excel(filepath)
    colslist = list(set(df[cols].values))

    copyfile(filepath, newfile)
    for j in colslist:
        writer = pd.ExcelWriter(newfile, engine='openpyxl')
        for myname in colslist:
            mydf = df.loc[df[cols] == myname]
            mydf.to_excel(writer, sheet_name=myname, index=False)
        writer.save()

    messagebox.showinfo('Output',
                        'You data has been split into {} and {} sheets has been created under single file named {new}.\n Click on OK .'.format(
                            ', '.join(colslist), len(colslist), new=newfile))

    label_head7 = Label(LogGui,
                        text='{n}The Files have been Splitted to different sheets.'.format(
                            n=now.strftime('%y-%m-%d %H:%M:%S')),
                        bd=1, relief='solid',
                        font='Times 10', anchor=N)
    label_head7.pack()

    print('\nCompleted')
    print('Thanks for using this program.')
    return


def Combine_GSTR2A_File2():
    import pandas as pd
    from tkinter import Label,Button,Entry
    import os
    import glob
    from tkinter import messagebox, filedialog
    import datetime
    import warnings
    import numpy as np

    FotaGui = Tk()

    LogGui = Tk()

    FotaGui.geometry("500x500")
    LogGui.geometry("250x250")

    FotaGui.title("Utility for Merging GSTR2A")
    LogGui.title("Log of all activities")

    Label_1 = Label(FotaGui, text="This is the utility for merging GSTR 2A", font="Times 16")
    Label_1.pack()

    warnings.filterwarnings('ignore')

    def file_path():
        global filepath
        global label_head7
        filepath = StringVar()
        global now

        now = datetime.datetime.now()

        # Fetch the file path of the hex file browsed.
        if (filepath == ""):
            filepath = filedialog.askopenfilename(initialdir=os.getcwd(),
                                                  title="Select a file",
                                                  filetypes=[("All Files", "*.*"), ("Pdf Files", "*.pdf"),
                                                             ("Text Files", "*.txt"), ("Excel FIles", "*.xlsx")])
        else:
            filepath = filedialog.askopenfilename(initialdir=filepath,
                                                  title="Select a file",
                                                  filetypes=[("All Files", "*.*"), ("Pdf Files", "*.pdf"),
                                                             ("Text Files", "*.txt"), ("Excel FIles", "*.xlsx")])

        extension = os.path.splitext(filepath)[1]
        filename = os.path.splitext(filepath)[0]
        pth = os.path.dirname(filepath)
        files = glob.glob(os.path.join(pth, '*{ext}'.format(ext=extension)))

        for f in files:
            label_head7 = Label(LogGui, text='{n}The File {fil} have been selected.'.format(fil=f, n=now.strftime(
                '%y-%m-%d %H:%M:%S')), bd=1, relief='solid',
                                font='Times 10', anchor=N)
            label_head7.pack()

    def Combine_GSTR2A_File():
        import pandas as pd
        import glob
        import os
        global filepath

        pth = os.path.dirname(filepath)

        filenames = glob.glob(pth + "/*.xlsx")

        i = 0
        for file in filenames:
            i = i + 1

        if i < 1:
            print("Upload at least 2 files")
        elif i > 60:
            print("Maximum capacity is 60 files at a time")
        else:
            pass

        cum_size = 0

        for file in filenames:
            size = os.path.getsize(file)

            cum_size = cum_size + size

            if size > 31457280:
                print("Please upload a smaller file size. Maximum limit is 30 mb.")

            elif cum_size > 314572800:
                print("Combined File size for all the file is more than 300 mb. Please use smaller files")
                break
            else:
                pass

        # A. iterate through each file to append it one below the other

        # A.1 : This will iterate through the B2B file

        df2 = pd.DataFrame()

        for files in filenames:
            df = pd.read_excel(files, sheet_name=1)

            df1 = df.drop([0, 1, 2, 3, 4], axis=0)

            df1 = df1.dropna(how='all')

            df1['File_name'] = files

            df2 = df2.append(df1)

        # this is used for deleting all the rows which are totally blank

        df3 = df2

        # this is used for renaming the names of the columns

        df3.rename(columns={'Goods and Services Tax  - GSTR 2A': 'GSTIN_of_Supplier'}, inplace=True)
        df3.rename(columns={'Unnamed: 1': 'Legal_Name_Of Supplier'}, inplace=True)
        df3.rename(columns={'Unnamed: 2': 'Inv_CN_DN_Number_Original'}, inplace=True)
        df3.rename(columns={'Unnamed: 3': 'Inv_CN_DN_Type_Original'}, inplace=True)
        df3.rename(columns={'Unnamed: 4': 'Inv_CN_DN_Date_Original'}, inplace=True)
        df3.rename(columns={'Unnamed: 5': 'Inv_CN_DN_Value_Original'}, inplace=True)
        df3.rename(columns={'Unnamed: 6': 'Place_Of_Supply'}, inplace=True)
        df3.rename(columns={'Unnamed: 7': 'Supply_Attract_Reverse_Charge'}, inplace=True)
        df3.rename(columns={'Unnamed: 8': 'GST_Rate'}, inplace=True)
        df3.rename(columns={'Unnamed: 9': 'Taxable_Value_Rs'}, inplace=True)
        df3.rename(columns={'Unnamed: 10': 'IGST_Rs'}, inplace=True)
        df3.rename(columns={'Unnamed: 11': 'CGST_Rs'}, inplace=True)
        df3.rename(columns={'Unnamed: 12': 'SGST_Rs'}, inplace=True)
        df3.rename(columns={'Unnamed: 13': 'Cess'}, inplace=True)
        df3.rename(columns={'Unnamed: 14': 'GSTR_1_5_Filing_Status'}, inplace=True)
        df3.rename(columns={'Unnamed: 15': 'GSTR_1_5_Filing_Date'}, inplace=True)
        df3.rename(columns={'Unnamed: 16': 'GSTR_1_5_Filing_Period'}, inplace=True)
        df3.rename(columns={'Unnamed: 17': 'GSTR_3B_Filing_Status'}, inplace=True)
        df3.rename(columns={'Unnamed: 18': 'Amendment_made_if_any'}, inplace=True)
        df3.rename(columns={'Unnamed: 19': 'Tax_Period_in_which_Amended'}, inplace=True)
        df3.rename(columns={'Unnamed: 20': 'Effective_date_of_cancellation'}, inplace=True)
        df3.rename(columns={'Unnamed: 21': 'Source'}, inplace=True)
        df3.rename(columns={'Unnamed: 22': 'IRN'}, inplace=True)
        df3.rename(columns={'Unnamed: 23': 'IRN_Date'}, inplace=True)

        # here we will remove the rows, in which the invoice number has  a total
        filt = df3['Inv_CN_DN_Number_Original'].str.contains('Total', na=False)
        df3 = df3[~filt]

        df3['Inv_CN_DN_Date_Unique'] = df3['Inv_CN_DN_Date_Original'].str.replace("-", ".")
        df3['Total_tax'] = df3['IGST_Rs'] + df3['CGST_Rs'] + df3['SGST_Rs']
        df3['Unique_ID'] = df3['GSTIN_of_Supplier'] + "/" + df3['Inv_CN_DN_Number_Original'] + "/" + df3[
            'Inv_CN_DN_Date_Unique']

        df3['Sheet_Name'] = ("B2B")

        df3['PAN_Number'] = df3["GSTIN_of_Supplier"].apply(lambda x: x[2:12:1])

        df3 = df3.replace(np.nan, "", regex=True)

        label_head21 = Label(FotaGui, text='The B2B table is being combined... Please wait')

        # A.2 : This will iterate through the B2BA file

        df2 = pd.DataFrame()

        for files in filenames:
            df = pd.read_excel(files, sheet_name=2)

            df1 = df.drop([0, 1, 2, 3, 4, 5], axis=0)

            df1 = df1.dropna(how='all')

            df1['File_name'] = files

            df2 = df2.append(df1)

        # this is used for deleting all the rows which are totally blank
        df4 = df2.dropna(how='all')

        # this is used for renaming the names of the columns

        df4.rename(
            columns={
                '                                      Goods and Services Tax - GSTR-2A': 'Inv_CN_DN_Number_Original'},
            inplace=True)
        df4.rename(columns={'Unnamed: 1': 'Inv_CN_DN_Date_Original'}, inplace=True)
        df4.rename(columns={'Unnamed: 2': 'GSTIN_of_Supplier'}, inplace=True)
        df4.rename(columns={'Unnamed: 3': 'Legal_Name_Of Supplier'}, inplace=True)
        df4.rename(columns={'Unnamed: 4': 'Inv_CN_DN_Type_Revised'}, inplace=True)
        df4.rename(columns={'Unnamed: 5': 'Inv_CN_DN_Number_Revised'}, inplace=True)
        df4.rename(columns={'Unnamed: 6': 'Inv_CN_DN_Date_Revised'}, inplace=True)
        df4.rename(columns={'Unnamed: 7': 'Inv_CN_DN_Value_Revised'}, inplace=True)
        df4.rename(columns={'Unnamed: 8': 'Place_Of_Supply'}, inplace=True)
        df4.rename(columns={'Unnamed: 9': 'Supply_Attract_Reverse_Charge'}, inplace=True)
        df4.rename(columns={'Unnamed: 10': 'GST_Rate'}, inplace=True)
        df4.rename(columns={'Unnamed: 11': 'Taxable_Value_Rs'}, inplace=True)
        df4.rename(columns={'Unnamed: 12': 'IGST_Rs'}, inplace=True)
        df4.rename(columns={'Unnamed: 13': 'CGST_Rs'}, inplace=True)
        df4.rename(columns={'Unnamed: 14': 'SGST_Rs'}, inplace=True)
        df4.rename(columns={'Unnamed: 15': 'Cess'}, inplace=True)
        df4.rename(columns={'Unnamed: 16': 'GSTR_1_5_Filing_Status'}, inplace=True)
        df4.rename(columns={'Unnamed: 17': 'GSTR_1_5_Filing_Date'}, inplace=True)
        df4.rename(columns={'Unnamed: 18': 'GSTR_1_5_Filing_Period'}, inplace=True)
        df4.rename(columns={'Unnamed: 19': 'GSTR_3B_Filing_Status'}, inplace=True)
        df4.rename(columns={'Unnamed: 20': 'Effective_date_of_cancellation'}, inplace=True)
        df4.rename(columns={'Unnamed: 21': 'Amendment_made_if_any'}, inplace=True)
        df4.rename(columns={'Unnamed: 22': 'Original_tax_period_in_which_reported'}, inplace=True)

        # here we will remove the rows, in which the invoice number has  a total
        filt = df4['Inv_CN_DN_Number_Revised'].str.contains('Total', na=False)

        df4 = df4[~filt]

        df4['Inv_CN_DN_Date_Unique'] = df4['Inv_CN_DN_Date_Original'].str.replace("-", ".")
        df4['Total_tax'] = df4['IGST_Rs'] + df4['CGST_Rs'] + df4['SGST_Rs']
        df4['Unique_ID'] = df4['GSTIN_of_Supplier'] + "/" + df4['Inv_CN_DN_Number_Original'] + "/" + df4[
            'Inv_CN_DN_Date_Unique']
        df4["Inv_CN_DN_Date_Revised_Unique"] = df4['Inv_CN_DN_Date_Revised'].str.replace("-", ".")

        df4['Sheet_Name'] = ("B2BA")

        df4['PAN_Number'] = df4["GSTIN_of_Supplier"].apply(lambda x: x[2:12:1])

        df4 = df4.replace(np.nan, "", regex=True)

        # A.3 : This will iterate through the CDNR file

        df2 = pd.DataFrame()

        for files in filenames:
            df = pd.read_excel(files, sheet_name=3)

            df1 = df.drop([0, 1, 2, 3, 4], axis=0)

            df1 = df1.dropna(how='all')

            df1['File_name'] = files

            df2 = df2.append(df1)

        # this is used for deleting all the rows which are totally blank
        df5 = df2.dropna(how='all')

        # this is used for renaming the names of the columns

        df5.rename(
            columns={
                '                                             Goods and Services Tax - GSTR-2A': 'GSTIN_of_Supplier'},
            inplace=True)
        df5.rename(columns={'Unnamed: 1': 'Legal_Name_Of Supplier'}, inplace=True)
        df5.rename(columns={'Unnamed: 2': 'Credit_Debit_Note_Original'}, inplace=True)
        df5.rename(columns={'Unnamed: 3': 'Inv_CN_DN_Number_Original'}, inplace=True)
        df5.rename(columns={'Unnamed: 4': 'Inv_CN_DN_Type_Original'}, inplace=True)
        df5.rename(columns={'Unnamed: 5': 'Inv_CN_DN_Date_Original'}, inplace=True)
        df5.rename(columns={'Unnamed: 6': 'Inv_CN_DN_Value_Original'}, inplace=True)
        df5.rename(columns={'Unnamed: 7': 'Place_Of_Supply'}, inplace=True)
        df5.rename(columns={'Unnamed: 8': 'Supply_Attract_Reverse_Charge'}, inplace=True)
        df5.rename(columns={'Unnamed: 9': 'GST_Rate'}, inplace=True)
        df5.rename(columns={'Unnamed: 10': 'Taxable_Value_Rs'}, inplace=True)
        df5.rename(columns={'Unnamed: 11': 'IGST_Rs'}, inplace=True)
        df5.rename(columns={'Unnamed: 12': 'CGST_Rs'}, inplace=True)
        df5.rename(columns={'Unnamed: 13': 'SGST_Rs'}, inplace=True)
        df5.rename(columns={'Unnamed: 14': 'Cess'}, inplace=True)
        df5.rename(columns={'Unnamed: 15': 'GSTR_1_5_Filing_Status'}, inplace=True)
        df5.rename(columns={'Unnamed: 16': 'GSTR_1_5_Filing_Date'}, inplace=True)
        df5.rename(columns={'Unnamed: 17': 'GSTR_1_5_Filing_Period'}, inplace=True)
        df5.rename(columns={'Unnamed: 18': 'GSTR_3B_Filing_Status'}, inplace=True)
        df5.rename(columns={'Unnamed: 19': 'Amendment_made_if_any'}, inplace=True)
        df5.rename(columns={'Unnamed: 20': 'Tax_Period_in_which_Amended'}, inplace=True)
        df5.rename(columns={'Unnamed: 21': 'Effective_date_of_cancellation'}, inplace=True)
        df5.rename(columns={'Unnamed: 22': 'Source'}, inplace=True)
        df5.rename(columns={'Unnamed: 23': 'IRN'}, inplace=True)
        df5.rename(columns={'Unnamed: 24': 'IRN_Date'}, inplace=True)

        # here we will remove the rows, in which the invoice number has  a total
        filt = df5['Inv_CN_DN_Number_Original'].str.contains('Total', na=False)

        df5 = df5[~filt]

        df5['Inv_CN_DN_Date_Unique'] = df5['Inv_CN_DN_Date_Original'].str.replace("-", ".")
        df5['Total_tax'] = df5['IGST_Rs'] + df5['CGST_Rs'] + df5['SGST_Rs']
        df5['Unique_ID'] = df5['GSTIN_of_Supplier'] + "/" + df5['Inv_CN_DN_Number_Original'] + "/" + df5[
            'Inv_CN_DN_Date_Unique']

        df5['Sheet_Name'] = ("CDNR")

        df5['PAN_Number'] = df5["GSTIN_of_Supplier"].apply(lambda x: x[2:12:1])

        df5 = df5.replace(np.nan, "", regex=True)

        # A.2 : This will iterate through the CDNRA file

        df2 = pd.DataFrame()

        for files in filenames:
            df = pd.read_excel(files, sheet_name=4)

            df1 = df.drop([0, 1, 2, 3, 4, 5], axis=0)

            df1 = df1.dropna(how='all')

            df1['File_name'] = files

            df2 = df2.append(df1)

        # this is used for deleting all the rows which are totally blank
        df6 = df2.dropna(how='all')

        # this is used for renaming the names of the columns

        df6.rename(
            columns={'                             Goods and Services Tax - GSTR2A': 'Credit_Debit_Note_Original'},
            inplace=True)
        df6.rename(columns={'Unnamed: 1': 'Inv_CN_DN_Number_Original'}, inplace=True)
        df6.rename(columns={'Unnamed: 2': 'Inv_CN_DN_Date_Original'}, inplace=True)
        df6.rename(columns={'Unnamed: 3': 'GSTIN_of_Supplier'}, inplace=True)
        df6.rename(columns={'Unnamed: 4': 'Legal_Name_Of Supplier'}, inplace=True)
        df6.rename(columns={'Unnamed: 5': 'Credit_Debit_Note_Revised'}, inplace=True)
        df6.rename(columns={'Unnamed: 6': 'Inv_CN_DN_Number_Revised'}, inplace=True)
        df6.rename(columns={'Unnamed: 7': 'Inv_CN_DN_Type_Revised'}, inplace=True)
        df6.rename(columns={'Unnamed: 8': 'Inv_CN_DN_Date_Revised'}, inplace=True)
        df6.rename(columns={'Unnamed: 9': 'Inv_CN_DN_Value_Revised'}, inplace=True)
        df6.rename(columns={'Unnamed: 10': 'Place_Of_Supply'}, inplace=True)
        df6.rename(columns={'Unnamed: 11': 'Supply_Attract_Reverse_Charge'}, inplace=True)
        df6.rename(columns={'Unnamed: 12': 'GST_Rate'}, inplace=True)
        df6.rename(columns={'Unnamed: 13': 'Taxable_Value_Rs'}, inplace=True)
        df6.rename(columns={'Unnamed: 14': 'IGST_Rs'}, inplace=True)
        df6.rename(columns={'Unnamed: 15': 'CGST_Rs'}, inplace=True)
        df6.rename(columns={'Unnamed: 16': 'SGST_Rs'}, inplace=True)
        df6.rename(columns={'Unnamed: 17': 'Cess'}, inplace=True)
        df6.rename(columns={'Unnamed: 18': 'GSTR_1_5_Filing_Status'}, inplace=True)
        df6.rename(columns={'Unnamed: 19': 'GSTR_1_5_Filing_Date'}, inplace=True)
        df6.rename(columns={'Unnamed: 20': 'GSTR_1_5_Filing_Period'}, inplace=True)
        df6.rename(columns={'Unnamed: 21': 'GSTR_3B_Filing_Status'}, inplace=True)
        df6.rename(columns={'Unnamed: 22': 'Amendment_made_if_any'}, inplace=True)
        df6.rename(columns={'Unnamed: 23': 'Original_tax_period_in_which_reported'}, inplace=True)
        df6.rename(columns={'Unnamed: 24': 'Effective_date_of_cancellation'}, inplace=True)

        # here we will remove the rows, in which the invoice number has  a total
        filt = df6['Inv_CN_DN_Number_Revised'].str.contains('Total', na=False)

        df6 = df6[~filt]

        df6['Inv_CN_DN_Date_Unique'] = df6['Inv_CN_DN_Date_Original'].str.replace("-", ".")
        df6['Total_tax'] = df6['IGST_Rs'] + df6['CGST_Rs'] + df6['SGST_Rs']
        df6['Unique_ID'] = df6['GSTIN_of_Supplier'] + "/" + df6['Inv_CN_DN_Number_Original'] + "/" + df6[
            'Inv_CN_DN_Date_Unique']

        df6["Inv_CN_DN_Date_Revised_Unique"] = df6['Inv_CN_DN_Date_Revised'].str.replace("-", ".")

        df6['Sheet_Name'] = ("CDNRA")

        df6['PAN_Number'] = df6["GSTIN_of_Supplier"].apply(lambda x: x[2:12:1])

        df6 = df6.replace(np.nan, "", regex=True)

        # Making a combined sheet with all merged

        df8 = df3.append(df4)

        df9 = df8.append(df5)

        df10 = df9.append(df6)

        df10['PAN_Number'] = df10["GSTIN_of_Supplier"].apply(lambda x: x[2:12:1])

        df10 = df10.replace(np.nan, "", regex=True)

        df10["Ultimate_Unique"] = df10["Sheet_Name"] + "/" + df10["Supply_Attract_Reverse_Charge"] + df10[
            "GSTR_1_5_Filing_Status"] + "/" + df10["Unique_ID"]

        df10["PAN_3_Way_Key"] = np.where(df10["Sheet_Name"] == "B2BA",
                                         df10["PAN_Number"] + "/" + df10["Inv_CN_DN_Number_Revised"] + "/"
                                         + df10["Inv_CN_DN_Date_Revised_Unique"],
                                         df10["PAN_Number"] + "/" + df10["Inv_CN_DN_Number_Original"]
                                         + "/" + df10["Inv_CN_DN_Date_Unique"])

        df10["PAN_2_Way_Key_PAN_InvNo"] = np.where(df10["Sheet_Name"] == "B2BA",
                                                   df10["PAN_Number"] + "/" + df10["Inv_CN_DN_Number_Revised"]
                                                   , df10["PAN_Number"] + "/" + df10["Inv_CN_DN_Number_Original"])

        df10["PAN_2_Way_Key_PAN_InvDt"] = np.where(df10["Sheet_Name"] == "B2BA",
                                                   df10["PAN_Number"] + "/" + df10["Inv_CN_DN_Date_Revised_Unique"]
                                                   , df10["PAN_Number"] + "/" + df10["Inv_CN_DN_Date_Unique"])

        # maiking a sheet with person who did not file the GSTR 1

        df11 = df10[df10['GSTR_1_5_Filing_Status'] == "N"]

        df12 = df10[(df10['Supply_Attract_Reverse_Charge'] == "Y") & (df10['GSTR_1_5_Filing_Status'] == "Y")]

        df13 = df10[(df10['Supply_Attract_Reverse_Charge'] == "N") & (df10['GSTR_1_5_Filing_Status'] == "Y") & (
                df10['Total_tax'] < 1)]

        df14 = df10[(df10['Supply_Attract_Reverse_Charge'] == "N") & (df10['GSTR_1_5_Filing_Status'] == "Y") & (
                df10['Total_tax'] >= 1)]

        # saving the file with the name "Combined"

        extension = os.path.splitext(filepath)[1]
        filename = os.path.splitext(filepath)[0]
        pth = os.path.dirname(filepath)
        newfile = os.path.join(pth, filename + 'GSTR2A_all_combined' + extension)

        writer = pd.ExcelWriter(newfile, engine='openpyxl')

        df3.to_excel(writer, sheet_name="B2B")

        df4.to_excel(writer, sheet_name="B2BA")

        df5.to_excel(writer, sheet_name="CDNR")

        df6.to_excel(writer, sheet_name="CDNRA")

        titles = list(df10.columns)

        titles[0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], \
        titles[10], titles[11], titles[12], titles[13], titles[14], titles[15], titles[16], titles[17], titles[18], \
        titles[
            19], titles[20], titles[21], titles[22], titles[23], titles[24], titles[25], titles[26], titles[27], titles[
            28], \
        titles[29], titles[30], titles[31], titles[32], titles[33], titles[34], titles[35] = titles[24], titles[28], \
                                                                                             titles[
                                                                                                 0], titles[1], titles[
                                                                                                 2], titles[3], titles[
                                                                                                 4], titles[5], titles[
                                                                                                 6], titles[7], titles[
                                                                                                 8], titles[9], titles[
                                                                                                 10], titles[11], \
                                                                                             titles[
                                                                                                 12], titles[13], \
                                                                                             titles[
                                                                                                 26], titles[25], \
                                                                                             titles[
                                                                                                 21], titles[27], \
                                                                                             titles[
                                                                                                 14], titles[15], \
                                                                                             titles[
                                                                                                 16], titles[17], \
                                                                                             titles[
                                                                                                 18], titles[19], \
                                                                                             titles[
                                                                                                 20], titles[22], \
                                                                                             titles[
                                                                                                 23], titles[29], \
                                                                                             titles[
                                                                                                 30], titles[31], \
                                                                                             titles[
                                                                                                 32], titles[33], \
                                                                                             titles[
                                                                                                 34], titles[35]

        df10[titles].to_excel(writer, sheet_name="All_Combined")

        titles = list(df11.columns)

        titles[0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], \
        titles[10], titles[11], titles[12], titles[13], titles[14], titles[15], titles[16], titles[17], titles[18], \
        titles[
            19], titles[20], titles[21], titles[22], titles[23], titles[24], titles[25], titles[26], titles[27], titles[
            28], \
        titles[29], titles[30], titles[31], titles[32], titles[33], titles[34], titles[35] = titles[24], titles[28], \
                                                                                             titles[
                                                                                                 0], titles[1], titles[
                                                                                                 2], titles[3], titles[
                                                                                                 4], titles[5], titles[
                                                                                                 6], titles[7], titles[
                                                                                                 8], titles[9], titles[
                                                                                                 10], titles[11], \
                                                                                             titles[
                                                                                                 12], titles[13], \
                                                                                             titles[
                                                                                                 26], titles[25], \
                                                                                             titles[
                                                                                                 21], titles[27], \
                                                                                             titles[
                                                                                                 14], titles[15], \
                                                                                             titles[
                                                                                                 16], titles[17], \
                                                                                             titles[
                                                                                                 18], titles[19], \
                                                                                             titles[
                                                                                                 20], titles[22], \
                                                                                             titles[
                                                                                                 23], titles[29], \
                                                                                             titles[
                                                                                                 30], titles[31], \
                                                                                             titles[
                                                                                                 32], titles[33], \
                                                                                             titles[
                                                                                                 34], titles[35]

        df11[titles].to_excel(writer, sheet_name="GSTR_1_Not Filed")

        titles = list(df12.columns)

        titles[0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], \
        titles[10], titles[11], titles[12], titles[13], titles[14], titles[15], titles[16], titles[17], titles[18], \
        titles[
            19], titles[20], titles[21], titles[22], titles[23], titles[24], titles[25], titles[26], titles[27], titles[
            28], \
        titles[29], titles[30], titles[31], titles[32], titles[33], titles[34], titles[35] = titles[24], titles[28], \
                                                                                             titles[
                                                                                                 0], titles[1], titles[
                                                                                                 2], titles[3], titles[
                                                                                                 4], titles[5], titles[
                                                                                                 6], titles[7], titles[
                                                                                                 8], titles[9], titles[
                                                                                                 10], titles[11], \
                                                                                             titles[
                                                                                                 12], titles[13], \
                                                                                             titles[
                                                                                                 26], titles[25], \
                                                                                             titles[
                                                                                                 21], titles[27], \
                                                                                             titles[
                                                                                                 14], titles[15], \
                                                                                             titles[
                                                                                                 16], titles[17], \
                                                                                             titles[
                                                                                                 18], titles[19], \
                                                                                             titles[
                                                                                                 20], titles[22], \
                                                                                             titles[
                                                                                                 23], titles[29], \
                                                                                             titles[
                                                                                                 30], titles[31], \
                                                                                             titles[
                                                                                                 32], titles[33], \
                                                                                             titles[
                                                                                                 34], titles[35]

        df12[titles].to_excel(writer, sheet_name="GSTR_Filed_RCM_Yes")

        titles = list(df13.columns)

        titles[0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], \
        titles[10], titles[11], titles[12], titles[13], titles[14], titles[15], titles[16], titles[17], titles[18], \
        titles[
            19], titles[20], titles[21], titles[22], titles[23], titles[24], titles[25], titles[26], titles[27], titles[
            28], \
        titles[29], titles[30], titles[31], titles[32], titles[33], titles[34], titles[35] = titles[24], titles[28], \
                                                                                             titles[
                                                                                                 0], titles[1], titles[
                                                                                                 2], titles[3], titles[
                                                                                                 4], titles[5], titles[
                                                                                                 6], titles[7], titles[
                                                                                                 8], titles[9], titles[
                                                                                                 10], titles[11], \
                                                                                             titles[
                                                                                                 12], titles[13], \
                                                                                             titles[
                                                                                                 26], titles[25], \
                                                                                             titles[
                                                                                                 21], titles[27], \
                                                                                             titles[
                                                                                                 14], titles[15], \
                                                                                             titles[
                                                                                                 16], titles[17], \
                                                                                             titles[
                                                                                                 18], titles[19], \
                                                                                             titles[
                                                                                                 20], titles[22], \
                                                                                             titles[
                                                                                                 23], titles[29], \
                                                                                             titles[
                                                                                                 30], titles[31], \
                                                                                             titles[
                                                                                                 32], titles[33], \
                                                                                             titles[
                                                                                                 34], titles[35]

        df13[titles].to_excel(writer, sheet_name="Tax_Zero_Cases")

        titles = list(df14.columns)

        titles[0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], \
        titles[10], titles[11], titles[12], titles[13], titles[14], titles[15], titles[16], titles[17], titles[18], \
        titles[
            19], titles[20], titles[21], titles[22], titles[23], titles[24], titles[25], titles[26], titles[27], titles[
            28], \
        titles[29], titles[30], titles[31], titles[32], titles[33], titles[34], titles[35] = titles[24], titles[28], \
                                                                                             titles[
                                                                                                 0], titles[1], titles[
                                                                                                 2], titles[3], titles[
                                                                                                 4], titles[5], titles[
                                                                                                 6], titles[7], titles[
                                                                                                 8], titles[9], titles[
                                                                                                 10], titles[11], \
                                                                                             titles[
                                                                                                 12], titles[13], \
                                                                                             titles[
                                                                                                 26], titles[25], \
                                                                                             titles[
                                                                                                 21], titles[27], \
                                                                                             titles[
                                                                                                 14], titles[15], \
                                                                                             titles[
                                                                                                 16], titles[17], \
                                                                                             titles[
                                                                                                 18], titles[19], \
                                                                                             titles[
                                                                                                 20], titles[22], \
                                                                                             titles[
                                                                                                 23], titles[29], \
                                                                                             titles[
                                                                                                 30], titles[31], \
                                                                                             titles[
                                                                                                 32], titles[33], \
                                                                                             titles[
                                                                                                 34], titles[35]

        df14[titles].to_excel(writer, sheet_name="Working_Cases")

        writer.save()

        messagebox.showinfo('Output', 'All GSTR2A files have been combined!. \n Click on OK')

    label_0 = Label(FotaGui, text='\n')
    label_0.pack()

    label_0 = Label(FotaGui, text='Step: 1 Select the File by clicking Browse Button !!!', font='Times 11', anchor=N,
                    bd=1, relief='solid')
    label_0.pack()

    Browsebutton = Button(FotaGui, width=15, text="BROWSE", command=file_path)
    Browsebutton.pack()

    label_head3 = Label(FotaGui, text='\n'
                                      '\n'
                        )
    label_head3.pack()

    label_head4 = Label(FotaGui, text='Step 2: Click on the action button at the below:', bd=1, relief='solid',
                        font='Times 12', anchor=N)

    label_head4.pack()

    Button_1 = Button(FotaGui, text="Combine GSTR2A files", command=Combine_GSTR2A_File)
    Button_1.pack()

    label_head12 = Label(FotaGui, text="   \n"
                                       "\n"
                                       "\n"
                                       "\n Feedback for improving the Program is sought."
                                       "\n Based on Feedback, program can be improved to Cater needs of Specific Users"
                                       "\n Send your feedback at pranav.tulshyan@gmail.com ", font="Times 10 ")

    label_head12.pack()

    LogGui.mainloop()

    FotaGui.mainloop()

def Clear_Memory():
    messagebox.showinfo('Memory Clear', 'The file selected have been cleared from memory')
    now = datetime.datetime.now()
    label_head12 = Label(LogGui,
                         text='{n}:The file selected have been cleared from memory. You may browse file again '.format(
                             n=now.strftime("%y-%m-%d %H:%M:%S")))
    label_head12.pack()


def reco_itr_2a():
    import numpy as np
    import openpyxl
    global fullpath
    global filepath

    global files_itr
    global files_con2a
    global files_mon2a

    # print(f'The MOn GSTR2A file path is {files_mon2a}')
    print(f'The Con GSTR2A file path is {files_con2a[0]}')
    print(f'The ITR file path is {files_itr[0]}')

    pth = os.path.dirname(str(files_con2a[0]))

    fullpath1 = pth + "/" + "Workings.xlsx"

    writer = pd.ExcelWriter(fullpath1, engine='xlsxwriter', options={'strings_to_formulas': True})
    #


    # writer.save()

    fullpath1a = pth + "/" + "Summary.xlsx"
    writer1 = pd.ExcelWriter(fullpath1a, engine='xlsxwriter', options={'strings_to_formulas': True})

    df1 = pd.DataFrame()
    df1.to_excel(writer1, sheet_name="Summary", index=False)

    writer1.save()

    fullpath2 = fullpath1a.replace("/", "\\")  # this is a very useful command for defining the correct filepath

    wb = load_workbook(fullpath2)
    ws = wb["Summary"]

    ws["B2"].value = "SUMMARY OF THE RECONCILIATION OF GSTR2A Vs ITR"
    ws.merge_cells("B2:F2")
    ws["C4"].value = "GSTR2A"
    ws.merge_cells("C4:D4")
    ws["E4"].value = "Purchase Register"
    ws.merge_cells("E4:F4")

    ws["B4"].value = "Particulars"
    ws.merge_cells("B4:B5")
    ws["C5"].value = "Count"
    ws["D5"].value = "Tax Amount"
    ws["E5"].value = "Count"
    ws["F5"].value = "Tax Amount"
    ws["B7"].value = "Total To be Matched"
    ws["B9"].value = "Matched with GST_INVNO_INVDATE_3_WAY"
    ws["B10"].value = "Matched with GST_INVNO_2_WAY"
    ws["B11"].value = "Matched with GST_INVDATE_2_WAY"

    ws["B13"].value = "Identified Possible Matches"

    ws["B15"].value = "Matched with PAN_INVNO_INVDATE_3_WAY"
    ws["B16"].value = "Matched with PAN_INVNO_2_WAY"
    ws["B17"].value = "Matched with PAN_INVDATE_2_WAY"

    ws["B19"].value = "Unmatched Cases"

    ws["B21"].value = "Check"

    # setting the tolerance limit for matching in Rupees

    tol_limit = int(150)

    ws["F1"].value = f"Tolerance Limit was {tol_limit}"

    gstr2a = pd.read_excel(files_con2a[0], sheet_name="Main_2A_Format",dtype={"Inv_CN_DN_Number_Final":str, "Inv_CN_DN_Date_Text":str, "Total_Tax":int})

    try:
        gstr2a['Inv_CN_DN_Number_Finall'] = gstr2a["Inv_CN_DN_Number_Final"].apply(lambda x: x.lower(str()))
    except:
        gstr2a['Inv_CN_DN_Number_Finall'] = gstr2a["Inv_CN_DN_Number_Final"]

    gstr2a['GST_INVNO_INVDATE_3_WAY'] = gstr2a['GSTIN_of_Supplier'] + "/" + gstr2a['Inv_CN_DN_Number_Finall'] + "/" + gstr2a['Inv_CN_DN_Date_Text']

    gstr2a['GST_INVNO_2_WAY'] = gstr2a['GSTIN_of_Supplier'] + "/" + gstr2a['Inv_CN_DN_Number_Finall']

    gstr2a['GST_INVDATE_2_WAY'] = gstr2a['GSTIN_of_Supplier'] + "/" + gstr2a['Inv_CN_DN_Date_Text']

    try:
        gstr2a['PAN_Number'] = gstr2a["GSTIN_of_Supplier"].apply(lambda x: x[2:12:1])
    except:
        gstr2a['PAN_Number'] = gstr2a["GSTIN_of_Supplier"]

    # the PAN number matches will be used as possible matches

    gstr2a['PAN_INVNO_INVDATE_3_WAY'] = gstr2a['PAN_Number'] + "/" + gstr2a['Inv_CN_DN_Number_Finall'] + "/" + \
                                        gstr2a['Inv_CN_DN_Date_Text']

    gstr2a['PAN_INVNO_2_WAY'] = gstr2a['PAN_Number'] + "/" + gstr2a['Inv_CN_DN_Number_Finall']

    gstr2a['PAN_INVDATE_2_WAY'] = gstr2a['PAN_Number'] + "/" + gstr2a['Inv_CN_DN_Date_Text']


    itr = pd.read_excel(files_itr[0], sheet_name="Main_ITR_Format",dtype={"Invoice_Number":str, "Invoice_Date_Text":str,"Total_Tax":int})

    try:
        itr["Invoice_Numberl"] = itr["Invoice_Number"].apply(lambda x: x.lower(str()))
    except:
        itr["Invoice_Numberl"] = itr["Invoice_Number"]

    itr["GST_INVNO_INVDATE_3_WAY"] = itr["Vendor_GST_REG"] + "/" + itr["Invoice_Numberl"] + "/" + itr[
        "Invoice_Date_Text"]

    itr["GST_INVNO_2_WAY"] = itr["Vendor_GST_REG"] + "/" + itr["Invoice_Numberl"]

    itr["GST_INVDATE_2_WAY"] = itr["Vendor_GST_REG"] + "/" + itr["Invoice_Date_Text"]

    try:
        itr["PAN_Number"] = itr["Vendor_GST_REG"].apply(lambda x: x[2:12:1])
    except:
        itr["PAN_Number"] = itr["Vendor_GST_REG"]

    # the PAN number matches will be used as possible matches

    itr["PAN_INVNO_INVDATE_3_WAY"] = itr["PAN_Number"] + "/" + itr["Invoice_Numberl"] + "/" + itr["Invoice_Date_Text"]

    itr["PAN_INVNO_2_WAY"] = itr["PAN_Number"] + "/" + itr["Invoice_Numberl"]

    itr["PAN_INVDATE_2_WAY"] = itr["PAN_Number"] + "/" + itr["Invoice_Date_Text"]

    ws["C7"].value = list(gstr2a.shape)[0]
    ws["D7"].value = sum(gstr2a["Total_Tax"])
    ws["E7"].value = list(itr.shape)[0]
    ws["F7"].value = sum(itr["Total_Tax"])

    # First Cut Matching : Here we will try to do that Matching based on 3 way i.e GST No, Inv No & Inv Date being same in ITR & GSTR2A

    gstr2a_pivot = pd.pivot_table(gstr2a, values="Total_Tax", index=["GST_INVNO_INVDATE_3_WAY"], aggfunc=np.sum)

    itr_pivot = pd.pivot_table(itr, values="Total_Tax", index=["GST_INVNO_INVDATE_3_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={'Total_Tax': 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={'Total_Tax': 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="GST_INVNO_INVDATE_3_WAY", right_on="GST_INVNO_INVDATE_3_WAY",
                                  how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    conditions = [compared["Difference_in_Tax"] > (tol_limit),

                  compared["Difference_in_Tax"] < (tol_limit * -1),

                  ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",

               "Exact Match within Tolerance"]

    compared["Remarks"] = np.select(conditions, results)

    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks"].values == "Exact Match within Tolerance"

    match_3_way_list = compared[mask_1]["GST_INVNO_INVDATE_3_WAY"].values

    mask_1a = gstr2a["GST_INVNO_INVDATE_3_WAY"].isin(match_3_way_list)  # returns a Boolean Array

    mask_1b = itr["GST_INVNO_INVDATE_3_WAY"].isin(match_3_way_list)  # returns a boolean array

    matched_gstr2a_3way = gstr2a[mask_1a]
    matched_gstr2a_3way["Matching Category"] = "3 Way matching GST + Inv No+ Inv Date"
    matched_itr_3way = itr[mask_1b]
    matched_itr_3way["Matching Category"] = "3 Way matching GST + Inv No+ Inv Date"

    ws["C9"].value = len(matched_gstr2a_3way["GST_INVNO_INVDATE_3_WAY"])
    ws["D9"].value = sum(matched_gstr2a_3way["Total_Tax"])
    ws["E9"].value = len(matched_itr_3way["GST_INVNO_INVDATE_3_WAY"])
    ws["F9"].value = sum(matched_itr_3way["Total_Tax"])

    bal_gstr2a_1cut = gstr2a[~mask_1a]
    bal_itr_1cut = itr[~mask_1b]

    # Second Cut Matching : Here we will try to do that Matching based on 2 way i.e GST No & Inv No

    gstr2a_pivot = pd.pivot_table(bal_gstr2a_1cut, values="Total_Tax", index=["GST_INVNO_2_WAY"], aggfunc=np.sum)

    itr_pivot = pd.pivot_table(bal_itr_1cut, values="Total_Tax", index=["GST_INVNO_2_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={'Total_Tax': 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={'Total_Tax': 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="GST_INVNO_2_WAY", right_on="GST_INVNO_2_WAY", how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    conditions = [compared["Difference_in_Tax"] > (tol_limit),

                  compared["Difference_in_Tax"] < (tol_limit * -1),

                  ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",

               "Exact Match within Tolerance"]

    compared["Remarks"] = np.select(conditions, results)

    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks"].values == "Exact Match within Tolerance"

    match_2_way_list1 = compared[mask_1]["GST_INVNO_2_WAY"].values

    mask_1a = bal_gstr2a_1cut["GST_INVNO_2_WAY"].isin(match_2_way_list1)  # returns a Boolean Array

    mask_1b = bal_itr_1cut["GST_INVNO_2_WAY"].isin(match_2_way_list1)  # returns a boolean array

    matched_gstr2a_2way1 = bal_gstr2a_1cut[mask_1a]
    matched_itr_2way1 = bal_itr_1cut[mask_1b]

    matched_gstr2a_2way1["Matching Category"] = "2 Way matching GST + Inv No"
    matched_itr_2way1["Matching Category"] = "2 Way matching GST + Inv No"

    ws["C10"].value = len(matched_gstr2a_2way1["GST_INVNO_2_WAY"])
    ws["D10"].value = sum(matched_gstr2a_2way1["Total_Tax"])
    ws["E10"].value = len(matched_itr_2way1["GST_INVNO_2_WAY"])
    ws["F10"].value = sum(matched_itr_2way1["Total_Tax"])

    bal_gstr2a_2cut = bal_gstr2a_1cut[~mask_1a]
    bal_itr_2cut = bal_itr_1cut[~mask_1b]

    # Third Cut Matching : Here we will try to do that Matching based on 2 way i.e GST No & Inv Date

    gstr2a_pivot = pd.pivot_table(bal_gstr2a_2cut, values="Total_Tax", index=["GST_INVDATE_2_WAY"], aggfunc=np.sum)

    itr_pivot = pd.pivot_table(bal_itr_2cut, values="Total_Tax", index=["GST_INVDATE_2_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={'Total_Tax': 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={'Total_Tax': 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="GST_INVDATE_2_WAY", right_on="GST_INVDATE_2_WAY", how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    conditions = [compared["Difference_in_Tax"] > (tol_limit),

                  compared["Difference_in_Tax"] < (tol_limit * -1),

                  ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",

               "Exact Match within Tolerance"]

    compared["Remarks"] = np.select(conditions, results)

    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks"].values == "Exact Match within Tolerance"

    match_2_way_list2 = compared[mask_1]["GST_INVDATE_2_WAY"].values

    mask_1a = bal_gstr2a_2cut["GST_INVDATE_2_WAY"].isin(match_2_way_list2)  # returns a Boolean Array

    mask_1b = bal_itr_2cut["GST_INVDATE_2_WAY"].isin(match_2_way_list2)  # returns a boolean array

    matched_gstr2a_2way2 = bal_gstr2a_2cut[mask_1a]
    matched_itr_2way2 = bal_itr_2cut[mask_1b]

    matched_gstr2a_2way2["Matching Category"] = "2 Way matching GST + Inv Date"
    matched_itr_2way2["Matching Category"] = "2 Way matching GST + Inv Date"

    ws["C11"].value = len(matched_gstr2a_2way2["GST_INVDATE_2_WAY"])
    ws["D11"].value = sum(matched_gstr2a_2way2["Total_Tax"])
    ws["E11"].value = len(matched_itr_2way2["GST_INVDATE_2_WAY"])
    ws["F11"].value = sum(matched_itr_2way2["Total_Tax"])

    bal_gstr2a_3cut = bal_gstr2a_2cut[~mask_1a]
    bal_itr_3cut = bal_itr_2cut[~mask_1b]

    #after the 3 cut matching, now we try to find out the Possible matches in form of PAN matching and upper /lower case matching
    # Fourth Cut Matching : Here we will try to do that Matching based on 3 way With PAN No Inv NO and Inv Date

    gstr2a_pivot = pd.pivot_table(bal_gstr2a_3cut, values="Total_Tax", index=["PAN_INVNO_INVDATE_3_WAY"], aggfunc=np.sum)

    itr_pivot = pd.pivot_table(bal_itr_3cut, values="Total_Tax", index=["PAN_INVNO_INVDATE_3_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={'Total_Tax': 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={'Total_Tax': 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="PAN_INVNO_INVDATE_3_WAY", right_on="PAN_INVNO_INVDATE_3_WAY", how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    conditions = [compared["Difference_in_Tax"] > (tol_limit),

                  compared["Difference_in_Tax"] < (tol_limit * -1),

                  ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",

               "Exact Match within Tolerance"]

    compared["Remarks"] = np.select(conditions, results)

    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks"].values == "Exact Match within Tolerance"

    match_3_way_list2 = compared[mask_1]["PAN_INVNO_INVDATE_3_WAY"].values

    mask_1a = bal_gstr2a_3cut["PAN_INVNO_INVDATE_3_WAY"].isin(match_3_way_list2)  # returns a Boolean Array

    mask_1b = bal_itr_3cut["PAN_INVNO_INVDATE_3_WAY"].isin(match_3_way_list2)  # returns a boolean array

    matched_gstr2a_3way2 = bal_gstr2a_3cut[mask_1a]
    matched_itr_3way2 = bal_itr_3cut[mask_1b]

    matched_gstr2a_3way2["Matching Category"] = "3 Way matching PAN + Inv No+ Inv Date"
    matched_itr_3way2["Matching Category"] = "3 Way matching PAN + Inv No + Inv Date"

    ws["C15"].value = len(matched_gstr2a_3way2["PAN_INVNO_INVDATE_3_WAY"])
    ws["D15"].value = sum(matched_gstr2a_3way2["Total_Tax"])
    ws["E15"].value = len(matched_itr_3way2["PAN_INVNO_INVDATE_3_WAY"])
    ws["F15"].value = sum(matched_itr_3way2["Total_Tax"])

    bal_gstr2a_4cut = bal_gstr2a_3cut[~mask_1a]
    bal_itr_4cut = bal_itr_3cut[~mask_1b]

    # Fifth Cut Matching : Here we will try to do that Matching based on 3 way With PAN No Inv NO and Inv Date

    gstr2a_pivot = pd.pivot_table(bal_gstr2a_4cut, values="Total_Tax", index=["PAN_INVNO_2_WAY"],
                                  aggfunc=np.sum)

    itr_pivot = pd.pivot_table(bal_itr_4cut, values="Total_Tax", index=["PAN_INVNO_2_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={'Total_Tax': 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={'Total_Tax': 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="PAN_INVNO_2_WAY", right_on="PAN_INVNO_2_WAY",
                                  how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    conditions = [compared["Difference_in_Tax"] > (tol_limit),

                  compared["Difference_in_Tax"] < (tol_limit * -1),

                  ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",

               "Exact Match within Tolerance"]

    compared["Remarks"] = np.select(conditions, results)

    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks"].values == "Exact Match within Tolerance"

    match_2_way_list3 = compared[mask_1]["PAN_INVNO_2_WAY"].values

    mask_1a = bal_gstr2a_4cut["PAN_INVNO_2_WAY"].isin(match_2_way_list3)  # returns a Boolean Array

    mask_1b = bal_itr_4cut["PAN_INVNO_2_WAY"].isin(match_2_way_list3)  # returns a boolean array

    matched_gstr2a_2way3 = bal_gstr2a_4cut[mask_1a]
    matched_itr_2way3 = bal_itr_4cut[mask_1b]

    matched_gstr2a_2way3["Matching Category"] = "2 Way matching PAN + Inv No"
    matched_itr_2way3["Matching Category"] = "2 Way matching PAN + Inv No "

    ws["C16"].value = len(matched_gstr2a_2way3["PAN_INVNO_2_WAY"])
    ws["D16"].value = sum(matched_gstr2a_2way3["Total_Tax"])
    ws["E16"].value = len(matched_itr_2way3["PAN_INVNO_2_WAY"])
    ws["F16"].value = sum(matched_itr_2way3["Total_Tax"])

    bal_gstr2a_5cut = bal_gstr2a_4cut[~mask_1a]
    bal_itr_5cut = bal_itr_4cut[~mask_1b]



    # Sixth Cut Matching : Here we will try to do that Matching based on 3 way With PAN No Inv NO and Inv Date

    gstr2a_pivot = pd.pivot_table(bal_gstr2a_5cut, values="Total_Tax", index=["PAN_INVDATE_2_WAY"],
                                  aggfunc=np.sum)

    itr_pivot = pd.pivot_table(bal_itr_4cut, values="Total_Tax", index=["PAN_INVDATE_2_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={'Total_Tax': 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={'Total_Tax': 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="PAN_INVDATE_2_WAY", right_on="PAN_INVDATE_2_WAY",
                                  how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    conditions = [compared["Difference_in_Tax"] > (tol_limit),

                  compared["Difference_in_Tax"] < (tol_limit * -1),

                  ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",

               "Exact Match within Tolerance"]

    compared["Remarks"] = np.select(conditions, results)

    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks"].values == "Exact Match within Tolerance"

    match_2_way_list4 = compared[mask_1]["PAN_INVDATE_2_WAY"].values

    mask_1a = bal_gstr2a_5cut["PAN_INVDATE_2_WAY"].isin(match_2_way_list4)  # returns a Boolean Array

    mask_1b = bal_itr_5cut["PAN_INVDATE_2_WAY"].isin(match_2_way_list4)  # returns a boolean array

    matched_gstr2a_2way4 = bal_gstr2a_5cut[mask_1a]
    matched_itr_2way4 = bal_itr_5cut[mask_1b]

    matched_gstr2a_2way4["Matching Category"] = "2 Way matching PAN + Inv Date"
    matched_itr_2way4["Matching Category"] = "2 Way matching PAN + Inv Date "

    ws["C17"].value = len(matched_gstr2a_2way4["PAN_INVDATE_2_WAY"])
    ws["D17"].value = sum(matched_gstr2a_2way4["Total_Tax"])
    ws["E17"].value = len(matched_itr_2way4["PAN_INVDATE_2_WAY"])
    ws["F17"].value = sum(matched_itr_2way4["Total_Tax"])

    bal_gstr2a_6cut = bal_gstr2a_5cut[~mask_1a]
    bal_itr_6cut = bal_itr_5cut[~mask_1b]

    gstr2a.to_excel(writer, sheet_name='Orignal GSTR2A', index=False)

    itr.to_excel(writer, sheet_name='Original ITR', index=False)

    all_matched_2a = pd.concat([matched_gstr2a_3way, matched_gstr2a_2way1, matched_gstr2a_2way2,matched_gstr2a_3way2, matched_gstr2a_2way3,matched_gstr2a_2way4], ignore_index=True)

    all_matched_itr = pd.concat([matched_itr_3way, matched_itr_2way1, matched_itr_2way2,matched_itr_3way2,matched_itr_2way3, matched_itr_2way4], ignore_index=True)

    all_matched_2a.to_excel(writer, sheet_name='Matched_GSTR2A', index=False)

    all_matched_itr.to_excel(writer, sheet_name='Matched_ITR', index=False)

    bal_gstr2a_6cut.to_excel(writer, sheet_name='Unmatched_GSTR2A', index=False)

    bal_itr_6cut.to_excel(writer, sheet_name='Unmatched_ITR', index=False)

    ws["C19"].value = len(bal_gstr2a_6cut["GST_INVDATE_2_WAY"])
    ws["D19"].value = sum(bal_gstr2a_6cut["Total_Tax"])
    ws["E19"].value = len(bal_itr_6cut["GST_INVDATE_2_WAY"])
    ws["F19"].value = sum(bal_itr_6cut["Total_Tax"])

    #     itr_only.to_excel(writer, sheet_name='Only in ITR', index=True)

    #     df = pd.DataFrame()

    #     df.to_excel(writer, sheet_name='Based on Pan_Inv Date>>>', index=True)

    writer.save()

    print("Success")

    wb.save(fullpath2)
    writer.save()


    messagebox.showinfo('Output', f'Matching has been done and saved in below path \n {fullpath2}\n Click on OK')

def download():
    import pandas as pd
    import numpy as np
    import openpyxl
    global fullpath
    global filepath

    global files_itr
    global files_con2a
    global files_mon2a


    pth = os.getcwd()

    fullpath1 = pth + "\\" + "Formats.xlsx"
    print(fullpath1)

    writer = pd.ExcelWriter(fullpath1, engine='xlsxwriter', options={'strings_to_formulas': True})

    dict1 = {"Vendor_GST_REG": ["Mandatory"], "Vendor_Name": ["Optional"], "Invoice_Number": ["Mandatory"],
             "Invoice_Date_Text": ["Mandatory"], "Total_Tax": ["Mandatory"], "IGST": ["Optional"], "SGST": ["Optional"],
             "CGST": ["Optional"], "UTGST": ["Optional"],"User Defined1":["Optional"],"User Defined2":["Optional"],"User Defined3":["Optional"],"User Defined4":["Optional"]}

    df1 = pd.DataFrame(dict1)
    df1.to_excel(writer, sheet_name="Main_ITR_Format", index=False)

    dict2={"GSTIN_of_Supplier":["Mandatory"],"Inv_CN_DN_Number_Final":["Mandatory"],"Legal_Name_Of Supplier":["Optional"],"Inv_CN_DN_Date_Text":["Mandatory"],"Total_Tax":["Mandatory"],"IGST":["Optional"],"SGST":["Optional"],"CGST":["Optional"],"UTGST":["Optional"],"User Defined1":["Optional"],"User Defined2":["Optional"],"User Defined3":["Optional"],"User Defined4":["Optional"]}
    df2 = pd.DataFrame(dict2)
    df2.to_excel(writer, sheet_name="Main_2A_Format", index=False)

    writer.save()
    messagebox.showinfo('Output', f'The Formats have been saved in below path \n {fullpath1}\n Click on OK')


# Below is the arrangement of the Text and the Button in Tkinter.


label_head0 = Label(FotaGui, text="   GSTR2A Reco for Corporate Users:"
                                  "\n Program for reconciling the GSTR2A and Purchase Register", font="Times 11 ")

label_head0.pack()

label_head0 = Label(FotaGui, text="\n Click on HELP_INFO for more Information", font="Times 11", anchor=W)
label_head0.pack()

Browsebutton = Button(FotaGui, width=15, text="HELP_INFO", command=HELP_INFO)
Browsebutton.pack()



label_0 = Label(FotaGui, text='\n If you have monthly GSTR2A file and need to combine \n Click on Combine GSTR2A Button'
                              '\n'
                , font='Times 12', anchor=N)
label_0.pack()

Browsebutton = Button(FotaGui, width=20, text="Combine GSTR2A Files", command=Combine_GSTR2A_File2)
Browsebutton.pack()


label_0 = Label(FotaGui, text='\n Download the Format of GSTR2A and the Purchase Register',
                font='Times 12', bd=1, relief='solid', anchor=N)
label_0.pack()

Browsebutton = Button(FotaGui, width=15, text="Download Format", command=download)
Browsebutton.pack()

label_0 = Label(FotaGui, text='\n If you already have the merged GSTR2A and Purchase Register , \n Follow the below Simple 2 steps for doing the GSTR2A reconcilitaion!!!',
                font='Times 12', bd=1, relief='solid', anchor=N)
label_0.pack()

label_0 = Label(FotaGui, text='\n Step: 1 Select the Combined GSTR2A Excel File by clicking Browse GSTR2A Button !!!',
                font='Times 12', anchor=N)
label_0.pack()

Browsebutton = Button(FotaGui, width=15, text="BROWSE GSTR2A", command=file_path_con2a)
Browsebutton.pack()

label_0 = Label(FotaGui, text='\n Step: 2 Select the Purchase Register File by clicking Browse ITR Button !!!',
                font='Times 12', anchor=N)
label_0.pack()

Browsebutton = Button(FotaGui, width=15, text="BROWSE ITR", command=file_path_itr)
Browsebutton.pack()


label_20 = Label(FotaGui, text=
'\n Step: 3 Click onReconcile Button to do the Reconciliation!!!'
'\n', font='Times 12', anchor=N)
label_20.pack()

Browsebutton = Button(FotaGui, width=20, text="Reconcile GSTR2A vs ITR", command=reco_itr_2a)
Browsebutton.pack()

# label_0 = Label(FotaGui, text='\n To split the Unmatched Purchase Register GSTN Wise, click on Split Files !!!',
#                 font='Times 12', anchor=N)
# label_0.pack()

#
# Browsebutton = Button(FotaGui, width=30, text="Split Unmatched Purchase Register", command=SPLIT_FILE)
# Browsebutton.pack()

label_head11 = Label(LogGui, text='Log of all Activities:', anchor=W)
label_head11.pack()

label_head12 = Label(FotaGui, text=
"\n Feedback for improving the Program is sought."
"\n Based on Feedback, program can be improved to Cater needs of Specific Users"
"\n Send your feedback at efficientcorporates.info@gmail.com ", bd=1, relief='solid',font="Times 10 ")

label_head12.pack()

LogGui.mainloop()

FotaGui.mainloop()
